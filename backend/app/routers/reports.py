"""报表管理
- 摘要: 总项目数/总金额/各状态计数
- 按月: 项目数量 / 金额月度趋势
- 按合作模式/单位/中标状态: 分组
- Excel 导出: 按筛选条件导出当前表格
权限：与项目同 (normal 仅自己, important 自己+下属, admin 全部)
"""
import io
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from app.auth import get_current_user
from app.database import get_db
from app.models import AIModelConfig, Project, ProjectFollowup, User, UserRole
from app.schemas import (
    AIAnalysisRequest, AIAnalysisResponse,
    AIReportAssistantRequest, AIReportAssistantResponse,
    FOLLOWUP_STAGE_CHOICES,
)

router = APIRouter(prefix="/api/reports", tags=["报表管理"])
log = logging.getLogger("reports")

WIN_BID_LABEL = {"yes": "中标", "no": "未中标", "in_progress": "进行中"}
STATUS_LABEL = {"pending_submit": "待提交", "pending_approval": "待审批", "approved": "已通过", "rejected": "已驳回"}
COOP_LABEL = {"long_term": "长期合作", "short_term": "短期合作"}
FEE_LABEL = {"mutual": "互免", "charged": "收费"}
SM_LABEL = {"yes": "是", "no": "否"}
FIELD_LABELS = {
    "source_label": "项目来源",
    "project_name": "项目名称",
    "project_code": "项目编号",
    "project_type": "项目类型",
    "partner_company": "合作公司名称",
    "responsible_sales": "责任销售",
    "project_amount": "项目金额",
    "expected_amount": "预计金额",
    "fee_amount": "费用金额",
    "approval_status": "审批状态",
    "win_bid_status": "中标状态",
    "tender_time": "招标日期",
    "bid_time": "投标日期",
    "latest_followup_stage": "最新跟单阶段",
    "latest_followup_progress": "最新跟单进展",
    "latest_followup_expected_amount": "最新跟单预计金额",
    "created_at": "创建时间",
    "updated_at": "更新时间",
}


def _scoped_query(db: Session, current_user: User):
    q = db.query(Project)
    if current_user.role == UserRole.normal:
        q = q.filter(Project.created_by == current_user.id)
    elif current_user.role == UserRole.important:
        child_ids = [c.id for c in current_user.children]
        child_ids.append(current_user.id)
        q = q.filter(or_(Project.created_by.in_(child_ids), Project.approver_id == current_user.id))
    return q


def _apply_filters(
    q,
    keyword: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
):
    if keyword:
        q = q.filter(or_(
            Project.project_name.contains(keyword),
            Project.partner_company.contains(keyword),
            Project.project_code.contains(keyword),
        ))
    if project_type:
        q = q.filter(Project.project_type == project_type)
    if project_name:
        q = q.filter(Project.project_name.contains(project_name))
    if responsible_sales:
        q = q.filter(Project.responsible_sales.contains(responsible_sales))
    if win_bid_status:
        q = q.filter(Project.win_bid_status == win_bid_status)
    if partner_company:
        q = q.filter(Project.partner_company.contains(partner_company))
    if amount_min is not None:
        q = q.filter(Project.project_amount >= amount_min)
    if amount_max is not None:
        q = q.filter(Project.project_amount <= amount_max)
    if start_date:
        try:
            q = q.filter(Project.tender_time >= date.fromisoformat(start_date))
        except Exception:
            pass
    if end_date:
        try:
            q = q.filter(Project.tender_time <= date.fromisoformat(end_date))
        except Exception:
            pass
    return q


def _latest_followup(project: Project):
    if not getattr(project, "followups", None):
        return None
    return max(project.followups, key=lambda item: item.created_at or datetime.min)


def _serialize_project_for_ai(project: Project, fields: Optional[list] = None):
    latest = _latest_followup(project)
    row = {
        "id": project.id,
        "source": project.source or "channel",
        "source_label": "自营项目" if project.source == "self" else "渠道项目",
        "project_name": project.project_name,
        "project_code": project.project_code,
        "project_type": project.project_type.value if hasattr(project.project_type, "value") else project.project_type,
        "partner_company": project.partner_company,
        "responsible_sales": project.responsible_sales,
        "project_amount": float(project.project_amount or 0),
        "expected_amount": float(project.expected_amount or 0),
        "fee_amount": float(project.fee_amount or 0),
        "approval_status": STATUS_LABEL.get(project.approval_status.value if hasattr(project.approval_status, "value") else project.approval_status, ""),
        "win_bid_status": WIN_BID_LABEL.get(project.win_bid_status.value if hasattr(project.win_bid_status, "value") else project.win_bid_status, ""),
        "tender_time": project.tender_time.isoformat() if project.tender_time else None,
        "bid_time": project.bid_time.isoformat() if project.bid_time else None,
        "created_at": project.created_at.isoformat() if project.created_at else None,
        "updated_at": project.updated_at.isoformat() if project.updated_at else None,
        "latest_followup_stage": latest.stage.value if latest and hasattr(latest.stage, "value") else (str(latest.stage) if latest else None),
        "latest_followup_progress": latest.progress if latest else None,
        "latest_followup_expected_amount": float(latest.expected_amount or 0) if latest else 0,
    }
    if fields:
        return {key: row.get(key) for key in fields if key in row}
    return row


def _resolve_model_info(db: Session, model_id: Optional[int]):
    if model_id:
        model = db.query(AIModelConfig).filter(AIModelConfig.id == model_id, AIModelConfig.is_enabled == True).first()
        if not model:
            raise HTTPException(status_code=404, detail="AI 模型配置不存在或未启用")
    else:
        model = db.query(AIModelConfig).filter(AIModelConfig.is_enabled == True).order_by(AIModelConfig.is_default.desc(), AIModelConfig.id.asc()).first()
    if not model:
        return None, None
    model_info = {
        "id": model.id,
        "name": model.name,
        "model_type": model.model_type,
        "provider": model.provider,
        "base_url": model.base_url,
        "model_name": model.model_name,
        "api_key": model.api_key,  # 服务端调用需要真 api_key；不要返回给前端
        "temperature": model.temperature,
        "max_tokens": model.max_tokens,
        "timeout_seconds": model.timeout_seconds,
        "is_enabled": model.is_enabled,
        "is_default": model.is_default,
        "notes": model.notes,
        "created_by": model.created_by,
        "created_at": model.created_at,
        "updated_at": model.updated_at,
        "creator": model.creator,
    }
    return model, model_info


# 系统提示词：优先用调用方传入，否则取数据库中 role_key='default' 且启用的那条
DEFAULT_SYSTEM_PROMPT = (
    "你是小销，销售项目数据智能助理。"
    "回答要简洁、可执行；遇到不确定的数据，请提示用户。"
)


def _resolve_system_prompt(db: Session, system_prompt: Optional[str]) -> tuple[str, str]:
    """返回 (system_prompt 文本, 显示用的角色名称)"""
    text = (system_prompt or "").strip()
    if text:
        # 推断角色名：取首行关键词
        first_line = text.splitlines()[0].strip().strip('"').strip('「').strip('」')
        role_label = first_line[:24] if first_line else "AI 助理"
        return text, role_label
    # fallback 到 DB
    try:
        from app.models import AgentPrompt
        row = (db.query(AgentPrompt)
               .filter(AgentPrompt.role_key == 'default', AgentPrompt.enabled == True)
               .order_by(AgentPrompt.id.desc()).first())
        if row and row.content:
            first_line = row.content.splitlines()[0].strip().strip('"').strip('「').strip('」')
            role_label = first_line[:24] if first_line else (row.name or "AI 助理")
            return row.content, role_label
    except Exception:
        pass
    return DEFAULT_SYSTEM_PROMPT, "小销"


def _build_report_snapshot(rows: list[Project]):
    total_rows = len(rows)
    total_amount = round(sum(float(item.project_amount or 0) for item in rows), 2)
    approved_count = sum(1 for item in rows if (item.approval_status.value if hasattr(item.approval_status, "value") else item.approval_status) == "approved")
    win_count = sum(1 for item in rows if (item.win_bid_status.value if hasattr(item.win_bid_status, "value") else item.win_bid_status) == "yes")
    self_count = sum(1 for item in rows if item.source == "self")
    channel_count = total_rows - self_count

    type_counter = {}
    partner_counter = {}
    sales_counter = {}
    followup_counter = {}
    for item in rows:
        project_type = item.project_type.value if hasattr(item.project_type, "value") else str(item.project_type or "未分类")
        type_counter[project_type] = type_counter.get(project_type, 0) + 1
        if item.partner_company:
            partner_counter[item.partner_company] = partner_counter.get(item.partner_company, 0) + 1
        if item.responsible_sales:
            sales_counter[item.responsible_sales] = sales_counter.get(item.responsible_sales, 0) + 1
        latest = _latest_followup(item)
        if latest:
            stage = latest.stage.value if hasattr(latest.stage, "value") else str(latest.stage)
            followup_counter[stage] = followup_counter.get(stage, 0) + 1

    top_type = max(type_counter.items(), key=lambda x: x[1])[0] if type_counter else "暂无"
    top_partner = max(partner_counter.items(), key=lambda x: x[1])[0] if partner_counter else "暂无"
    top_sales = max(sales_counter.items(), key=lambda x: x[1])[0] if sales_counter else "暂无"
    top_stage = max(followup_counter.items(), key=lambda x: x[1])[0] if followup_counter else "暂无"
    approval_rate = round((approved_count / total_rows * 100), 1) if total_rows else 0
    win_rate = round((win_count / total_rows * 100), 1) if total_rows else 0
    summary_text = (
        f"当前筛选范围共 {total_rows} 个项目，项目总金额 {total_amount:,.2f}。"
        f"其中自营项目 {self_count} 个、渠道项目 {channel_count} 个；审批通过率 {approval_rate}%，中标率 {win_rate}%。"
        f"项目类型以“{top_type}”最多，合作公司以“{top_partner}”最集中，责任销售以“{top_sales}”最多，跟单阶段主要集中在“{top_stage}”。"
    )
    return {
        "total_rows": total_rows,
        "total_amount": total_amount,
        "approved_count": approved_count,
        "win_count": win_count,
        "self_count": self_count,
        "channel_count": channel_count,
        "type_counter": type_counter,
        "partner_counter": partner_counter,
        "sales_counter": sales_counter,
        "followup_counter": followup_counter,
        "summary_text": summary_text,
    }


def _call_llm(model_info: dict, system_prompt: str, user_prompt: str, history: list | None = None) -> str:
    """调用 OpenAI 兼容接口的大模型，支持本地（Ollama/vLLM/LM Studio）和云端（Kimi/MiniMax/DeepSeek）。
    返回模型回复的纯文本。任何异常都会被上层 try/except 兜住，并 fallback 到骨架回答。
    """
    try:
        from openai import OpenAI  # 延迟导入，未装包时降级走骨架
    except Exception:
        log.warning("openai 包未安装，无法调用大模型；返回 None 让上层 fallback")
        return None
    if not model_info:
        return None
    api_key = (model_info.get('api_key') or '').strip() or 'sk-no-key'
    base_url = (model_info.get('base_url') or '').strip() or None
    model_name = model_info.get('model_name') or ''
    if not model_name:
        return None
    timeout = int(model_info.get('timeout_seconds') or 60)
    temperature = float(model_info.get('temperature') or 0.2)
    max_tokens = model_info.get('max_tokens')

    # OpenAI Python SDK 不会自动拼 /v1，必须把 base_url 归一为带 /v1 的根路径
    sdk_base_url = base_url.rstrip('/') if base_url else None
    if sdk_base_url:
        if sdk_base_url.endswith('/chat/completions'):
            sdk_base_url = sdk_base_url[: -len('/chat/completions')].rstrip('/')
        if not sdk_base_url.endswith('/v1'):
            sdk_base_url = sdk_base_url + '/v1'

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    # 历史对话（最近 6 条，避免 token 爆炸）
    for msg in (history or [])[-6:]:
        role = msg.get('role')
        content = (msg.get('content') or '').strip()
        if not content or role not in ('user', 'assistant'):
            continue
        messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": user_prompt})

    try:
        client_kwargs = {"api_key": api_key, "timeout": timeout}
        if sdk_base_url:
            client_kwargs["base_url"] = sdk_base_url
        client = OpenAI(**client_kwargs)
        kwargs = {
            "model": model_name,
            "messages": messages,
            "temperature": temperature,
            "stream": False,
        }
        if max_tokens:
            kwargs["max_tokens"] = int(max_tokens)
        log.info("LLM CALL -> base_url=%s (raw=%s) model=%s timeout=%s messages=%d", sdk_base_url, base_url, model_name, timeout, len(messages))
        resp = client.chat.completions.create(**kwargs)
        # 兼容字符串/对象两种返回
        if hasattr(resp, 'choices') and resp.choices:
            msg = resp.choices[0].message
            content = (msg.content or '') if hasattr(msg, 'content') else (msg.get('content') if isinstance(msg, dict) else '')
            content = (content or '').strip()
            log.info("LLM RESP <- choices=%d content_len=%d", len(resp.choices), len(content))
            return content or None
        log.warning("LLM RESP <- no choices, raw_type=%s raw_preview=%r",
                    type(resp).__name__,
                    str(resp)[:200] if not hasattr(resp, 'choices') else None)
        return None
    except Exception as e:
        import traceback
        log.warning("LLM 调用失败: %s\n%s", e, traceback.format_exc())
        return None


def _answer_as_xiaoxiao(question: str, snapshot: dict, role_label: str = "小销"):
    q = (question or "").strip()
    if not q:
        return snapshot["summary_text"]
    role_prefix = f"我是{role_label}。"
    if "中标" in q:
        return f"{role_prefix}当前范围内中标项目 {snapshot['win_count']} 个，中标率约 {round((snapshot['win_count'] / snapshot['total_rows'] * 100), 1) if snapshot['total_rows'] else 0}% 。如果你愿意，我下一步可以继续按项目类型或责任销售拆开看。"
    if "自营" in q or "渠道" in q:
        return f"{role_prefix}当前自营项目 {snapshot['self_count']} 个，渠道项目 {snapshot['channel_count']} 个。整体上 {'自营项目更多' if snapshot['self_count'] > snapshot['channel_count'] else '渠道项目更多' if snapshot['channel_count'] > snapshot['self_count'] else '两类项目数量接近'}。"
    if "合作" in q or "公司" in q or "客户" in q:
        top_partner = max(snapshot["partner_counter"].items(), key=lambda x: x[1])[0] if snapshot["partner_counter"] else "暂无"
        return f"{role_prefix}从合作公司分布看，当前最集中的合作公司是“{top_partner}”。如果你要，我可以继续按合作公司给你做一个 Top 排名说明。"
    if "销售" in q:
        top_sales = max(snapshot["sales_counter"].items(), key=lambda x: x[1])[0] if snapshot["sales_counter"] else "暂无"
        return f"{role_prefix}按责任销售看，当前项目数最多的是“{top_sales}”。如果你希望，我可以进一步说明不同责任销售负责的金额规模。"
    if "阶段" in q or "跟单" in q:
        top_stage = max(snapshot["followup_counter"].items(), key=lambda x: x[1])[0] if snapshot["followup_counter"] else "暂无"
        return f"{role_prefix}跟单阶段里，目前最集中的阶段是“{top_stage}”。这通常意味着这批项目主要还停留在这个推进阶段。"
    if "金额" in q or "多少" in q or "总额" in q:
        return f"{role_prefix}当前筛选范围的项目总金额是 {snapshot['total_amount']:,.2f}。如果你想更细一点，我可以继续按项目类型、责任销售或合作公司拆分金额。"
    return f"{role_prefix}{snapshot['summary_text']} 如果你继续追问某个方向，比如中标、金额、责任销售、合作公司或跟单阶段，我可以继续展开。"


def _append_current_table_row(ws, project: Project):
    project_type_val = project.project_type.value if hasattr(project.project_type, "value") else project.project_type
    ws.append([
        project.project_code or "",
        project.project_name or "",
        project_type_val or "",
        project.responsible_sales or "",
        project.partner_company or "",
        WIN_BID_LABEL.get(project.win_bid_status.value if hasattr(project.win_bid_status, "value") else project.win_bid_status, ""),
        float(project.project_amount or 0),
        STATUS_LABEL.get(project.approval_status.value if hasattr(project.approval_status, "value") else project.approval_status, ""),
        project.tender_time.isoformat() if project.tender_time else "",
        project.bid_time.isoformat() if project.bid_time else "",
        project.creator.real_name if project.creator else "",
        project.approver.real_name if project.approver else "",
        project.created_at.strftime("%Y-%m-%d %H:%M") if project.created_at else "",
    ])


def _append_full_row(ws, project: Project):
    latest = _latest_followup(project)
    ws.append([
        project.id,
        "自营项目" if project.source == "self" else "渠道项目",
        project.project_code or "",
        project.project_name or "",
        project.project_type.value if hasattr(project.project_type, "value") else project.project_type,
        project.form_instance_id or "",
        project.partner_company or "",
        project.company_address or "",
        project.owner_contact_person or "",
        project.owner_contact_info or "",
        project.main_qualification or "",
        project.legal_representative or "",
        project.contact_person or "",
        project.contact_info or "",
        COOP_LABEL.get(project.cooperation_mode.value if hasattr(project.cooperation_mode, "value") else project.cooperation_mode, ""),
        FEE_LABEL.get(project.fee_mode.value if hasattr(project.fee_mode, "value") else project.fee_mode, ""),
        float(project.project_amount or 0),
        float(project.expected_amount or 0),
        float(project.fee_amount or 0),
        SM_LABEL.get(project.is_sm.value if hasattr(project.is_sm, "value") else project.is_sm, ""),
        WIN_BID_LABEL.get(project.win_bid_status.value if hasattr(project.win_bid_status, "value") else project.win_bid_status, ""),
        STATUS_LABEL.get(project.approval_status.value if hasattr(project.approval_status, "value") else project.approval_status, ""),
        project.responsible_sales or "",
        project.project_overview or "",
        project.tender_time.isoformat() if project.tender_time else "",
        project.bid_time.isoformat() if project.bid_time else "",
        project.creator.real_name if project.creator else "",
        project.approver.real_name if project.approver else "",
        project.storage_zone.name if project.storage_zone else "",
        project.tender_folder or "",
        project.bid_folder or "",
        latest.stage.value if latest and hasattr(latest.stage, "value") else (str(latest.stage) if latest else ""),
        latest.progress if latest else "",
        float(latest.expected_amount or 0) if latest else 0,
        project.created_at.strftime("%Y-%m-%d %H:%M") if project.created_at else "",
        project.updated_at.strftime("%Y-%m-%d %H:%M") if project.updated_at else "",
    ])


@router.get("/summary")
def summary(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max)
    total = q.count()
    total_amount = q.with_entities(func.coalesce(func.sum(Project.project_amount), 0.0)).scalar() or 0.0
    fee_amount = q.with_entities(func.coalesce(func.sum(Project.fee_amount), 0.0)).scalar() or 0.0
    status_rows = q.with_entities(Project.approval_status, func.count(Project.id), func.coalesce(func.sum(Project.project_amount), 0.0)).group_by(Project.approval_status).all()
    by_status = [{"status": st.value if hasattr(st, "value") else str(st), "count": int(cnt), "amount": float(amt)} for st, cnt, amt in status_rows]
    return {"total": total, "total_amount": float(total_amount), "fee_amount": float(fee_amount), "by_status": by_status}


@router.get("/trend")
def trend(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max)
    month_expr = func.strftime("%Y-%m", Project.tender_time)
    rows = q.with_entities(month_expr.label("month"), func.count(Project.id), func.coalesce(func.sum(Project.project_amount), 0.0)).group_by("month").order_by("month").all()
    return [{"month": month, "count": int(count), "amount": float(amount)} for month, count, amount in rows if month]


@router.get("/by-partner")
def by_partner(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max)
    rows = q.with_entities(Project.partner_company, func.count(Project.id), func.coalesce(func.sum(Project.project_amount), 0.0)).group_by(Project.partner_company).order_by(func.count(Project.id).desc()).limit(limit).all()
    return [{"partner": partner, "count": int(count), "amount": float(amount)} for partner, count, amount in rows]


@router.get("/by-cooperation")
def by_cooperation(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max)
    rows = q.with_entities(Project.cooperation_mode, func.count(Project.id), func.coalesce(func.sum(Project.project_amount), 0.0)).group_by(Project.cooperation_mode).all()
    return [{
        "mode": mode.value if hasattr(mode, "value") else str(mode),
        "label": COOP_LABEL.get(mode.value if hasattr(mode, "value") else str(mode), str(mode)),
        "count": int(count),
        "amount": float(amount),
    } for mode, count, amount in rows]


@router.get("/by-win-bid")
def by_win_bid(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max)
    rows = q.with_entities(Project.win_bid_status, func.count(Project.id), func.coalesce(func.sum(Project.project_amount), 0.0)).group_by(Project.win_bid_status).all()
    return [{
        "status": status.value if hasattr(status, "value") else str(status),
        "label": WIN_BID_LABEL.get(status.value if hasattr(status, "value") else str(status), str(status)),
        "count": int(count),
        "amount": float(amount),
    } for status, count, amount in rows]


@router.get("/by-followup-stage")
def by_followup_stage(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    scoped = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max).subquery()
    visible_proj_ids = db.query(scoped.c.id).subquery()
    sub = db.query(ProjectFollowup.project_id, func.max(ProjectFollowup.created_at).label("mx")).filter(ProjectFollowup.project_id.in_(visible_proj_ids)).group_by(ProjectFollowup.project_id).subquery()
    latest = db.query(ProjectFollowup).join(sub, (ProjectFollowup.project_id == sub.c.project_id) & (ProjectFollowup.created_at == sub.c.mx)).all()
    stat = {stage: {"stage": stage, "count": 0, "expected_amount": 0.0} for stage in FOLLOWUP_STAGE_CHOICES}
    for item in latest:
        stage = item.stage.value if hasattr(item.stage, "value") else str(item.stage)
        if stage not in stat:
            stat[stage] = {"stage": stage, "count": 0, "expected_amount": 0.0}
        stat[stage]["count"] += 1
        stat[stage]["expected_amount"] += float(item.expected_amount or 0)
    result = []
    for stage in FOLLOWUP_STAGE_CHOICES:
        row = stat[stage]
        row["expected_amount"] = round(row["expected_amount"], 2)
        result.append(row)
    return result


@router.get("/export")
def export(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    project_name: Optional[str] = None,
    responsible_sales: Optional[str] = None,
    win_bid_status: Optional[str] = None,
    partner_company: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), keyword, start_date, end_date, project_type, project_name, responsible_sales, win_bid_status, partner_company, amount_min, amount_max).options(joinedload(Project.creator), joinedload(Project.approver)).order_by(Project.id.desc())
    rows = q.all()
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "当前筛选表格"
        ws.append(["项目编号", "项目名称", "项目类型", "责任销售", "合作公司名称", "中标状态", "项目金额(元)", "审批状态", "招标日期", "投标日期", "创建人", "审批人", "创建时间"])
        for project in rows:
            _append_current_table_row(ws, project)
        for col, width in enumerate([14, 30, 12, 14, 24, 12, 14, 10, 12, 12, 10, 10, 16], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"projects_selected_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={fname}"})
    except ImportError:
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["项目编号", "项目名称", "项目类型", "责任销售", "合作公司名称", "中标状态", "项目金额", "审批状态"])
        for project in rows:
            project_type_val = project.project_type.value if hasattr(project.project_type, "value") else project.project_type
            writer.writerow([
                project.project_code or "",
                project.project_name or "",
                project_type_val or "",
                project.responsible_sales or "",
                project.partner_company or "",
                WIN_BID_LABEL.get(project.win_bid_status.value if hasattr(project.win_bid_status, "value") else project.win_bid_status, ""),
                float(project.project_amount or 0),
                STATUS_LABEL.get(project.approval_status.value if hasattr(project.approval_status, "value") else project.approval_status, ""),
            ])
        data = buf.getvalue().encode("utf-8-sig")
        fname = f"projects_selected_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(io.BytesIO(data), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/export-full")
def export_full(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _scoped_query(db, current_user).options(joinedload(Project.creator), joinedload(Project.approver), joinedload(Project.storage_zone), joinedload(Project.followups)).order_by(Project.id.desc())
    rows = q.all()
    try:
        from openpyxl import Workbook
        wb = Workbook()
        channel_ws = wb.active
        channel_ws.title = "渠道项目"
        self_ws = wb.create_sheet("自营项目")
        headers = [
            "项目ID", "项目来源", "项目编号", "项目名称", "项目类型", "表单实例ID",
            "合作单位", "公司地址", "业主联系人", "业主联系方式", "主要资质", "法定代表",
            "联系人", "联系方式", "合作模式", "费用模式", "项目金额(元)", "预计金额(元)", "费用金额(元)",
            "是否SM", "中标状态", "审批状态", "责任销售", "项目概述",
            "招标日期", "投标日期", "创建人", "审批人", "存储区域", "招标资料目录", "投标文档目录",
            "最新跟单阶段", "最新跟单进展", "最新跟单预计金额(元)", "创建时间", "更新时间",
        ]
        for ws in [channel_ws, self_ws]:
            ws.append(headers)
        for project in rows:
            _append_full_row(self_ws if project.source == "self" else channel_ws, project)
        for ws in [channel_ws, self_ws]:
            for col, width in enumerate([10, 12, 16, 30, 12, 12, 22, 20, 12, 16, 18, 12, 12, 14, 12, 12, 14, 14, 14, 8, 10, 10, 12, 30, 12, 12, 10, 10, 12, 20, 20, 12, 30, 16, 16, 16], 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"projects_full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={fname}"})
    except ImportError:
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["项目ID", "项目来源", "项目编号", "项目名称", "项目类型", "合作单位", "项目金额", "预计金额", "费用金额", "审批状态"])
        for project in rows:
            writer.writerow([
                project.id,
                "自营项目" if project.source == "self" else "渠道项目",
                project.project_code,
                project.project_name,
                project.project_type.value if hasattr(project.project_type, "value") else project.project_type,
                project.partner_company,
                project.project_amount,
                project.expected_amount,
                project.fee_amount,
                project.approval_status.value if hasattr(project.approval_status, "value") else project.approval_status,
            ])
        data = buf.getvalue().encode("utf-8-sig")
        fname = f"projects_full_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(io.BytesIO(data), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/ai-analyze", response_model=AIAnalysisResponse)
def ai_analyze(
    data: AIAnalysisRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(_scoped_query(db, current_user), data.keyword, data.start_date, data.end_date, data.project_type, data.project_name, data.responsible_sales, data.win_bid_status, data.partner_company, data.amount_min, data.amount_max).options(joinedload(Project.followups)).order_by(Project.id.desc())
    rows = q.limit(20).all()
    all_rows = q.all()
    selected_fields = data.fields or ["source_label", "project_name", "project_type", "partner_company", "project_amount", "fee_amount", "approval_status", "latest_followup_stage"]
    _, model_info = _resolve_model_info(db, data.model_id)
    preview_rows = [_serialize_project_for_ai(item, selected_fields) for item in rows]
    snapshot = _build_report_snapshot(all_rows)
    system_prompt_text, role_label = _resolve_system_prompt(db, data.system_prompt)

    # 拼装给 LLM 用的 user prompt：业务上下文 + 用户要求
    ctx_lines = [
        f"## 当前数据范围",
        f"- 共匹配 {snapshot['total_rows']} 个项目",
        f"- 项目总金额: {snapshot['total_amount']:,.2f}",
        f"- 中标: {snapshot['win_count']} / 已审批通过: {snapshot['approved_count']}",
        f"- 自营: {snapshot['self_count']} / 渠道: {snapshot['channel_count']}",
        f"- 摘要: {snapshot['summary_text']}",
        "",
        f"## 已选字段（用于左侧表格渲染）",
        ", ".join(FIELD_LABELS.get(f, f) for f in selected_fields),
        "",
        f"## 预览样本（最多 20 条）",
    ]
    for i, item in enumerate(preview_rows, 1):
        ctx_lines.append(f"  {i}. " + " | ".join(
            f"{FIELD_LABELS.get(k, k)}={preview_rows[i-1][k]}" for k in selected_fields
        ))
    ctx_lines += ["", f"## 用户的要求", data.prompt or "请基于以上数据给出一句话总结。"]
    user_prompt = "\n".join(ctx_lines)

    # 真接 LLM；失败 fallback 到骨架答案
    llm_answer = _call_llm(model_info, system_prompt_text, user_prompt)
    if llm_answer:
        final_answer = llm_answer
        answer_mode = "llm"
    else:
        final_answer = _answer_as_xiaoxiao(data.prompt, snapshot, role_label)
        answer_mode = "skeleton"
    # 暴露给前端前脱敏 api_key
    if model_info:
        model_info["api_key"] = None

    return {
        "mode": answer_mode,
        "message": (
            f"已按当前筛选条件完成数据预览，当前共匹配 {snapshot['total_rows']} 条数据。"
            + ("（已由大模型生成摘要）" if answer_mode == "llm" else "（当前未连接大模型，使用骨架回答）")
        ),
        "model": model_info,
        "agent": {
            "name": role_label,
            "role": "智能助理",
            "abilities": ["数据摘要", "问答沟通", "趋势提醒"],
            "system_prompt": system_prompt_text,
        },
        "filters": {
            "keyword": data.keyword,
            "project_type": data.project_type,
            "project_name": data.project_name,
            "responsible_sales": data.responsible_sales,
            "win_bid_status": data.win_bid_status,
            "partner_company": data.partner_company,
            "amount_min": data.amount_min,
            "amount_max": data.amount_max,
            "start_date": data.start_date,
            "end_date": data.end_date,
        },
        "fields": selected_fields,
        "field_labels": {field: FIELD_LABELS.get(field, field) for field in selected_fields},
        "display_type": data.display_type,
        "total_rows": snapshot["total_rows"],
        "summary_text": snapshot["summary_text"],
        "answer": final_answer,
        "preview_rows": preview_rows,
        "suggestions": [
            "当前导出 Excel 会导出当前筛选表格，全量导出会按渠道项目和自营项目拆成两个 sheet。",
            "表格预览已改为中文表头，和系统里的业务字段名称保持一致。",
            f"{role_label}已加载系统提示词，所有回复都将按该角色风格生成。",
            "已根据用户要求把上下文 + 数据摘要发送给大模型，模型回复直接展示在 answer 中。",
        ],
    }


@router.post("/ai-assistant", response_model=AIReportAssistantResponse)
def ai_assistant(
    data: AIReportAssistantRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_filters(
        _scoped_query(db, current_user),
        data.keyword, data.start_date, data.end_date,
        data.project_type, data.project_name, data.responsible_sales,
        data.win_bid_status, data.partner_company, data.amount_min, data.amount_max,
    ).options(joinedload(Project.followups)).order_by(Project.id.desc())
    rows = q.all()
    _, model_info = _resolve_model_info(db, data.model_id)
    snapshot = _build_report_snapshot(rows)
    system_prompt_text, role_label = _resolve_system_prompt(db, data.system_prompt)

    # 把数据上下文 + 用户问题拼给 LLM
    ctx_lines = [
        f"## 当前数据范围（共 {snapshot['total_rows']} 个项目）",
        f"- 项目总金额: {snapshot['total_amount']:,.2f}",
        f"- 中标: {snapshot['win_count']} / 已审批通过: {snapshot['approved_count']}",
        f"- 自营: {snapshot['self_count']} / 渠道: {snapshot['channel_count']}",
        f"- 摘要: {snapshot['summary_text']}",
        "",
        f"## 用户的问题",
        data.question,
    ]
    user_prompt = "\n".join(ctx_lines)

    llm_answer = _call_llm(model_info, system_prompt_text, user_prompt, history=data.history)
    if llm_answer:
        final_answer = llm_answer
        answer_mode = "llm"
    else:
        final_answer = _answer_as_xiaoxiao(data.question, snapshot, role_label)
        answer_mode = "skeleton"
    # 暴露给前端前脱敏 api_key
    if model_info:
        model_info["api_key"] = None

    return {
        "assistant_name": role_label,
        "mode": answer_mode,
        "model": model_info,
        "total_rows": snapshot["total_rows"],
        "summary_text": snapshot["summary_text"],
        "answer": final_answer,
        "tips": [
            "你可以继续问：哪类项目最多？" if answer_mode == "skeleton" else "模型已基于你的提问 + 当前数据范围直接作答。",
            "你可以继续问：责任销售的金额分布怎么样？",
            "你可以继续问：中标项目和未中标项目差异是什么？",
            f"{role_label}已加载管理员配置的系统提示词，所有回答都按该角色风格生成。",
        ],
    }
